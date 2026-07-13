from io import BytesIO
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def build_rekap_excel(surat_list, q="", bulan="", tahun=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Rekapitulasi Surat"

    # =========================
    # TITLE
    # =========================
    ws.merge_cells("A1:G1")
    ws["A1"] = "REKAPITULASI PENGIRIMAN SURAT KE POS BERDASARKAN X5"
    ws["A1"].font = Font(name="Arial", size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:G2")

    if bulan and tahun:
        subtitle = f"BULAN {bulan.upper()} {tahun}"
    elif bulan:
        subtitle = f"BULAN {bulan.upper()}"
    elif tahun:
        subtitle = f"TAHUN {tahun}"
    else:
        subtitle = "SELURUH DATA REKAPITULASI"

    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Arial", size=12, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A3:G3")
    ws["A3"] = f"Tanggal Export: {datetime.now().strftime('%d-%m-%Y %H:%M')}"
    ws["A3"].font = Font(name="Arial", size=9, italic=True)
    ws["A3"].alignment = Alignment(horizontal="right", vertical="center")

    # =========================
    # HEADER
    # =========================
    headers = [
        "No.",
        "Bulan",
        "Tahun",
        "Nomor Referensi",
        "Kode Surat",
        "Klasifikasi",
        "Alamat",
    ]

    start_row = 5

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = header
        cell.font = Font(name="Arial", size=10, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = header_fill
        cell.border = border

    # =========================
    # DATA
    # =========================
    for row_idx, surat in enumerate(surat_list, start=start_row + 1):
        values = [
            surat["nomor_urut"],
            surat["bulan"],
            surat["tahun"],
            surat["nomor_referensi"],
            surat["kode_surat"],
            surat["klasifikasi"],
            surat["alamat"],
        ]

        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(
                horizontal="center" if col_idx in [1, 2, 3, 5] else "left",
                vertical="center",
                wrap_text=True
            )

    # =========================
    # LAYOUT
    # =========================
    column_widths = {
        "A": 8,
        "B": 14,
        "C": 10,
        "D": 30,
        "E": 14,
        "F": 26,
        "G": 70,
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[start_row].height = 28

    for row in range(start_row + 1, ws.max_row + 1):
        ws.row_dimensions[row].height = 55

    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:G{ws.max_row}"

    # Page setup
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    return bio