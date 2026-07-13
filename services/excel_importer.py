from openpyxl import load_workbook
import re


def normalize_text(value):
    if value is None:
        return ""

    text = str(value).strip()

    # Kalau nilai Excel error seperti #N/A, tetap jadikan string kosong
    # supaya tidak mengganggu proses import.
    if text.upper() in ["#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NAME?"]:
        return ""

    return text


def normalize_header(value):
    if value is None:
        return ""

    text = str(value).strip().lower()

    # Hilangkan spasi berlebih
    text = re.sub(r"\s+", " ", text)

    # Hilangkan tanda baca umum, contoh:
    # "No." menjadi "no"
    # "No. Surat" menjadi "no surat"
    text = text.replace(".", "")
    text = text.replace(":", "")
    text = text.replace("-", " ")

    text = re.sub(r"\s+", " ", text).strip()

    return text


def canonical_header(header):
    """
    Mengubah variasi nama header Excel menjadi nama standar.
    """

    header = normalize_header(header)

    aliases = {
        "no": "no",
        "nomor": "no",
        "nomor urut": "no",
        "no urut": "no",

        "nomor referensi": "nomor_referensi",
        "no referensi": "nomor_referensi",
        "no surat": "nomor_referensi",
        "nomor surat": "nomor_referensi",
        "nomor referensi surat": "nomor_referensi",

        "kode surat": "kode_surat",
        "kode": "kode_surat",

        "klasifikasi": "klasifikasi",
        "jenis surat": "klasifikasi",

        "alamat": "alamat",
        "ditujukan": "alamat",
        "tujuan": "alamat",
        "alamat tujuan": "alamat",
    }

    return aliases.get(header, "")


def find_header_row(ws):
    """
    Mencari baris header Excel.
    Minimal harus menemukan 3 dari header standar:
    no, nomor_referensi, kode_surat, klasifikasi, alamat
    """

    for row in ws.iter_rows(min_row=1, max_row=30):
        canonical_values = []

        for cell in row:
            canonical = canonical_header(cell.value)
            if canonical:
                canonical_values.append(canonical)

        unique_headers = set(canonical_values)

        if len(unique_headers) >= 3:
            return row[0].row

    return None


def get_header_map(ws, header_row):
    """
    Membuat mapping:
    nama_header_standar -> nomor_kolom_excel
    """

    header_map = {}

    for cell in ws[header_row]:
        canonical = canonical_header(cell.value)

        if canonical:
            header_map[canonical] = cell.column

    return header_map


def get_cell_value(ws, row_num, col_num):
    if not col_num:
        return ""

    return normalize_text(ws.cell(row=row_num, column=col_num).value)


def import_rekap_from_excel(file_stream, bulan, tahun):
    """
    Membaca file Excel dan mengubahnya menjadi list data surat.

    Kolom Excel yang didukung:
    - No / No. / Nomor Urut
    - Nomor referensi / No Surat / Nomor Surat
    - Kode Surat / Kode
    - Klasifikasi / Jenis Surat
    - Alamat / Ditujukan / Tujuan / Alamat Tujuan

    Bulan dan tahun diambil dari input form import.
    """

    wb = load_workbook(file_stream, data_only=True)
    ws = wb.active

    header_row = find_header_row(ws)

    if not header_row:
        raise ValueError(
            "Header Excel tidak ditemukan. Pastikan ada kolom: No, Nomor referensi, Kode Surat, Klasifikasi, dan Alamat."
        )

    header_map = get_header_map(ws, header_row)

    col_no = header_map.get("no")
    col_nomor_referensi = header_map.get("nomor_referensi")
    col_kode_surat = header_map.get("kode_surat")
    col_klasifikasi = header_map.get("klasifikasi")
    col_alamat = header_map.get("alamat")

    if not col_no or not col_nomor_referensi or not col_alamat:
        raise ValueError(
            "Kolom wajib tidak lengkap. Minimal harus ada: No, Nomor referensi, dan Alamat."
        )

    rows = []

    for row_num in range(header_row + 1, ws.max_row + 1):
        nomor_urut = get_cell_value(ws, row_num, col_no)
        nomor_referensi = get_cell_value(ws, row_num, col_nomor_referensi)
        kode_surat = get_cell_value(ws, row_num, col_kode_surat)
        klasifikasi = get_cell_value(ws, row_num, col_klasifikasi)
        alamat = get_cell_value(ws, row_num, col_alamat)

        # Skip baris kosong
        if not nomor_urut and not nomor_referensi and not alamat:
            continue

        # Skip kalau baris masih header duplikat
        if normalize_header(nomor_referensi) in [
            "nomor referensi",
            "no surat",
            "nomor surat"
        ]:
            continue

        rows.append({
            "bulan": bulan,
            "tahun": tahun,
            "nomor_urut": nomor_urut,
            "nomor_referensi": nomor_referensi,
            "kode_surat": kode_surat,
            "klasifikasi": klasifikasi,
            "alamat": alamat,
        })

    return rows